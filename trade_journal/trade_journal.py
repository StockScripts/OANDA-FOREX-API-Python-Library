import pandas as pd
import warnings
import numpy as np
import pdb
import logging
from apis.ser_data_obj import ser_data_obj
from trade_journal.trade import Trade
from trade_journal.trade_list import TradeList
from openpyxl import load_workbook, Workbook
from configparser import ConfigParser
from pivot.pivotlist import PivotList

# create logger
tj_logger = logging.getLogger(__name__)
tj_logger.setLevel(logging.INFO)

class TradeJournal(object):
    '''
    Constructor

    Class variables
    ---------------
    url: path to the .xlsx file with the trade journal
    worksheet: str, Required
               Name of the worksheet that will be used to create the object.
               i.e. trading_journal
    settingf : str, Optional
               Path to *.ini file with settings
    settings : ConfigParser object generated using 'settingf'
               Optional
    ser_data_f : str, Optional
                 ser_data file
    ser_data_obj : ser_data_obj, Optional
                   ser_data_obj with serialized data
    '''

    def __init__(self, url, worksheet, settingf=None, settings=None, ser_data_f=None,
                 ser_data_obj=None):
        self.url = url
        self.worksheet = worksheet
        self.settingf = settingf

        if ser_data_f is not None:
            self.ser_data_obj = ser_data_obj(ifile=ser_data_f)
        elif ser_data_obj is not None:
            self.ser_data_obj = ser_data_obj
        else:
            self.ser_data_obj = None

        #read-in the 'trading_journal' worksheet from a .xlsx file into a pandas dataframe
        try:
            xls_file = pd.ExcelFile(url)
            df = xls_file.parse(worksheet, converters={'start': str, 'end': str, 'trend_i': str})
            # replace n.a. string by NaN
            df = df.replace('n.a.', np.NaN)
            # remove trailing whitespaces from col names
            df.columns = df.columns.str.rstrip()
            self.df = df
        except FileNotFoundError:
            wb = Workbook()
            wb.create_sheet(worksheet)
            wb.save(str(self.url))

        if self.settingf is not None:
            # parse settings file (in .ini file)
            parser = ConfigParser()
            parser.read(settingf)
            self.settings = parser
        else:
            self.settings = settings

    def fetch_tradelist(self):
        '''
        This function will produce a TradeList object
        using self.url

        Returns
        -------
        TradeList
        '''

        trade_list = []
        for index, row in self.df.iterrows():
            pair = row['id'].split(" ")[0]
            args = {
                'pair': pair,
                'settingf': self.settingf
            }
            for c in row.keys():
                args[c] = row[c]
            t = Trade(**args)
            trade_list.append(t)

        tl = TradeList(settingf=self.settingf,
                       tlist=trade_list,
                       ser_data_obj=self.ser_data_obj)
        return tl

    def write_tradelist(self, trade_list):
        '''
        Write the TradeList to the Excel spreadsheet
        pointed by the trade_journal

        Parameters
        ----------
        tradelist : TradeList, Required

        Returns
        -------
        Nothing
        '''

        assert self.settings.has_option('trade_journal', 'worksheet_name'), \
            "'worksheet_name' needs to be defined"

        # get colnames to print in output worksheet from settings
        assert self.settings.has_option('trade_journal', 'colnames'), "'colnames' needs to be defined"
        colnames = self.settings.get('trade_journal', 'colnames').split(",")

        data = []
        for t in trade_list.tlist:
            row = []
            for a in colnames:
                value = None
                try:
                    value = getattr(t, a)
                    if isinstance(value, PivotList):
                        dt_l = value.print_pivots_dates()
                        value = [d.strftime('%d/%m/%Y:%H:%M') for d in dt_l]
                except:
                    tj_logger.warn("No value for attribute: {0}".format(a))
                    value = "n.a."
                row.append(value)
            data.append(row)
        df = pd.DataFrame(data, columns=colnames)

        book = load_workbook(self.url)
        writer = pd.ExcelWriter(self.url, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        tj_logger.info("Creating new worksheet with trades with name: {0}".
                       format(self.settings.get('trade_journal', 'worksheet_name')))
        df.to_excel(writer, self.settings.get('trade_journal', 'worksheet_name'))
        writer.save()


